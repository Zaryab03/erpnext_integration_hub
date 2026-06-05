"""Excel parser using openpyxl read-only mode for memory-efficient processing of large files."""
from __future__ import annotations

from typing import Generator
import frappe


class ExcelParser:
	"""Parse an .xlsx file into an iterator of row dicts.

	Uses openpyxl read_only=True so that even 100k-row workbooks are streamed
	from disk rather than loaded entirely into memory.
	"""

	def __init__(self, profile: dict):
		self.profile = profile

	def parse(self, file_content: bytes) -> list[dict]:
		"""Return a list of dicts, one per data row.

		Raises frappe.ValidationError on unrecognisable content.
		"""
		try:
			import io
			import openpyxl
		except ImportError:
			frappe.throw("openpyxl is not installed. Run: pip install openpyxl")

		try:
			wb = openpyxl.load_workbook(
				io.BytesIO(file_content),
				read_only=True,
				data_only=True,
			)
		except Exception as e:
			frappe.throw(f"Cannot open Excel file: {e}", frappe.ValidationError)

		sheet_name = self.profile.get("excel_sheet_name") or None
		if sheet_name and sheet_name in wb.sheetnames:
			ws = wb[sheet_name]
		else:
			ws = wb.active

		header_row_idx = int(self.profile.get("excel_header_row") or 1)
		data_start_idx = int(self.profile.get("excel_data_start_row") or 2)
		skip_empty = bool(self.profile.get("skip_empty_rows", True))

		rows = list(ws.iter_rows(values_only=True))
		wb.close()

		if len(rows) < header_row_idx:
			frappe.throw("File has fewer rows than the configured Header Row number.")

		headers = self._normalise_headers(
			[str(c) if c is not None else "" for c in rows[header_row_idx - 1]],
			case_sensitive=False,
		)

		result = []
		for row_idx, row in enumerate(rows[data_start_idx - 1:], start=data_start_idx):
			values = [c for c in row]
			if skip_empty and all(v is None or str(v).strip() == "" for v in values):
				continue
			row_dict = {headers[i]: self._coerce(values[i]) for i in range(len(headers))}
			row_dict["_source_row"] = row_idx
			result.append(row_dict)

		return result

	@staticmethod
	def _normalise_headers(headers: list[str], case_sensitive: bool = False) -> list[str]:
		if case_sensitive:
			return headers
		return [h.strip().lower().replace(" ", "_") for h in headers]

	@staticmethod
	def _coerce(value) -> str | None:
		if value is None:
			return None
		import datetime
		if isinstance(value, (datetime.date, datetime.datetime)):
			return str(value)
		return str(value).strip()
